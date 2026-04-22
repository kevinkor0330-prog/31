"""
load_sales_order.py
엑셀 영업실적 데이터를 읽어서
sales.sales_order 테이블에 INSERT합니다.
 
실행 전 준비:
    pip install openpyxl psycopg2-binary
 
실행:
    python load_sales_order.py
"""
 
import psycopg2
from openpyxl import load_workbook
from datetime import date
 
# ── DB 연결 정보 ───────────────────────────────────────────────
DB = {
    "host":     "158.247.208.195",
    "port":     5432,
    "dbname":   "db_hwaseong",
    "user":     "admin",
    "password": "1234",
}
EXCEL_PATH = "74A9CC40.xlsx"
# ─────────────────────────────────────────────────────────────
 
 
def load_partner_map(conn) -> dict[str, int]:
    """
    DB에서 { 거래처명: id } 딕셔너리를 가져옵니다.
    엑셀 거래처명 → partner_id 변환에 사용합니다.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT id, name FROM sales.res_partner")
        rows = cur.fetchall()
    return {name: pid for pid, name in rows}
 
 
def load_product_map(conn) -> dict[str, int]:
    """
    DB에서 { 제품군명: id } 딕셔너리를 가져옵니다.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT id, category FROM sales.product")
        rows = cur.fetchall()
    return {category: pid for pid, category in rows}
 
 
def parse_row(row, partner_map: dict, product_map: dict) -> dict | None:
    """
    엑셀 한 행을 받아서 DB INSERT용 딕셔너리로 변환합니다.
    변환 불가능한 행은 None 반환.
    """
 
    # ── 거래처명 정제 ──────────────────────────────────────────
    raw_name = row[2]
    if not raw_name:
        return None
    partner_name = raw_name.strip()
 
    partner_id = partner_map.get(partner_name)
    if partner_id is None:
        print(f"  ⚠ 거래처 없음 (스킵): {repr(partner_name)}")
        return None
 
    # ── 제품군 ────────────────────────────────────────────────
    product_category = row[4]
    product_id = product_map.get(product_category)  # 없으면 None (허용)
 
    # ── 기간: 년도(26) → 2026, 월 → DATE ─────────────────────
    year  = 2000 + int(row[0])
    month = int(row[1])
    period = date(year, month, 1)
 
    # ── 금액 ─────────────────────────────────────────────────
    def to_decimal(val):
        """None 이나 빈값이면 None, 아니면 float 변환"""
        if val is None or val == '':
            return None
        return float(val)
 
    supply_amount  = to_decimal(row[5])   # 공급가액
    freight_amount = to_decimal(row[6])   # 운임비 및 금형비 (149개가 None)
    tax_amount     = to_decimal(row[10])  # 부가세합계금액
    collected_amount = to_decimal(row[17]) # 회수금액
 
    # ── 수량 / 중량 ───────────────────────────────────────────
    quantity_ea = int(row[11]) if row[11] else None
    weight_ton  = float(row[8]) if row[8] else None
 
    # ── 채권 ─────────────────────────────────────────────────
    payment_type = row[13]  # '현금' or '어음'
 
    def to_date(val):
        """datetime → date, None이면 None"""
        if val is None:
            return None
        return val.date() if hasattr(val, 'date') else val
 
    invoice_date   = to_date(row[14])  # 세금계산서발행일자
    due_date       = to_date(row[15])  # 채권회수예정일자
    collected_date = to_date(row[16])  # 채권회수일자 (31개 None)
 
    # ── 미수확인 플래그 ───────────────────────────────────────
    # 엑셀에서 '미수' 라고 적힌 것만 True
    is_uncollected = (row[20] == '미수')
 
    # ── 기타 ─────────────────────────────────────────────────
    actual_freight = to_decimal(row[21])  # 실제운임비지급금액 (전부 None)
    note           = row[22] if row[22] else None
 
    return {
        "period":            period,
        "partner_id":        partner_id,
        "product_id":        product_id,
        "salesperson":       row[12],
        "supply_amount":     supply_amount,
        "freight_amount":    freight_amount,
        "tax_amount":        tax_amount,
        "quantity_ea":       quantity_ea,
        "weight_ton":        weight_ton,
        "payment_type":      payment_type,
        "invoice_date":      invoice_date,
        "due_date":          due_date,
        "collected_date":    collected_date,
        "collected_amount":  collected_amount,
        "is_uncollected":    is_uncollected,
        "actual_freight":    actual_freight,
        "note":              note,
    }
 
 
def insert_sales_orders(conn, records: list[dict]) -> int:
    """
    변환된 레코드 리스트를 DB에 INSERT합니다.
    성공한 건수를 반환합니다.
    """
    sql = """
        INSERT INTO sales.sales_order (
            period, partner_id, product_id, salesperson,
            supply_amount, freight_amount, tax_amount,
            quantity_ea, weight_ton,
            payment_type, invoice_date, due_date,
            collected_date, collected_amount,
            is_uncollected, actual_freight, note
        ) VALUES (
            %(period)s, %(partner_id)s, %(product_id)s, %(salesperson)s,
            %(supply_amount)s, %(freight_amount)s, %(tax_amount)s,
            %(quantity_ea)s, %(weight_ton)s,
            %(payment_type)s, %(invoice_date)s, %(due_date)s,
            %(collected_date)s, %(collected_amount)s,
            %(is_uncollected)s, %(actual_freight)s, %(note)s
        )
    """
    count = 0
    with conn.cursor() as cur:
        for rec in records:
            cur.execute(sql, rec)
            count += 1
    conn.commit()
    return count
 
 
def main():
    # ── 1. 엑셀 읽기 ──────────────────────────────────────────
    print(f"엑셀 읽는 중: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"총 {len(all_rows)}행 발견\n")
 
    # ── 2. DB 연결 & 매핑 로드 ────────────────────────────────
    print("DB 연결 중...")
    conn = psycopg2.connect(**DB)
 
    partner_map = load_partner_map(conn)
    product_map = load_product_map(conn)
    print(f"거래처 {len(partner_map)}개, 제품 {len(product_map)}개 로드 완료\n")
 
    # ── 3. 행 변환 ────────────────────────────────────────────
    print("데이터 변환 중...")
    records = []
    for row in all_rows:
        parsed = parse_row(row, partner_map, product_map)
        if parsed:
            records.append(parsed)
 
    print(f"변환 성공: {len(records)}건 / 전체 {len(all_rows)}건\n")
 
    # ── 4. INSERT ─────────────────────────────────────────────
    try:
        count = insert_sales_orders(conn, records)
        print(f"✓ INSERT 완료: {count}건")
    except Exception as e:
        conn.rollback()
        print(f"에러 발생, rollback: {e}")
        raise
    finally:
        conn.close()
 
 
if __name__ == "__main__":
    main()