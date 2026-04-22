"""
load_partners.py
엑셀의 거래처(대분류) 컬럼을 읽어서
sales.res_partner 테이블에 parent/child 구조로 INSERT합니다.
"""


import psycopg2
from openpyxl import load_workbook

DB = {
    "host":     "158.247.208.195",  # vultr 서버 IP
    "port":     5432,
    "dbname":   "db_hwaseong",
    "user":     "admin",
    "password": "1234",
}
EXCEL_PATH = "74A9CC40.xlsx"



 
def parse_partners(raw_names: set[str]) -> dict:
    """
    거래처명 집합을 받아서 parent/child 관계를 분석합니다.
 
    케이스 1 — 슬래시 있고 base가 단독 존재:
        '알파어셈블리솔루션즈코리아'  → parent (type=contact)
        '알파어셈블리솔루션즈코리아/시흥' → child  (type=delivery)
 
    케이스 2 — 슬래시 있지만 base가 단독으로 없음:
        '코웨이/유구공장' → parent '코웨이' 를 자동 생성 후 child 처리
        '도담축산/아이스팩' → 동일
 
    케이스 3 — 괄호 (가화텍(W1,W2), 가화텍(W3,W4)):
        슬래시와 동일 로직. '가화텍' 이 parent
 
    케이스 4 — 슬래시/괄호 없음:
        단독 거래처 → parent (type=contact)
 
    반환값:
        {
          'parents': { name: {} },          # 본사/메인 거래처
          'children': { name: parent_name } # 납품처 → 본사 이름
        }
    """
    parents = {}
    children = {}
 
    for name in raw_names:
        if '/' in name:
            base, child_label = name.split('/', 1)
            # base를 parent로 등록 (아직 없으면)
            if base not in parents:
                parents[base] = {}
            # 현재 name은 child
            children[name] = base
 
        elif '(' in name and ')' in name:
            base = name.split('(')[0].strip()
            if base not in parents:
                parents[base] = {}
            children[name] = base
 
        else:
            # 슬래시/괄호 없는 단독 거래처
            # 나중에 child로 판명되면 덮어쓰지 않게 children 체크 먼저
            if name not in children:
                parents[name] = {}
 
    # child로 분류된 이름이 parents에도 있으면 parents에서 제거
    # (예: '알파어셈블리솔루션즈코리아' 는 parent 로 유지, 문제없음)
    # 단, child 이름 자체가 parents에 있으면 제거
    for child_name in children:
        parents.pop(child_name, None)
 
    return {"parents": parents, "children": children}
 
 
def insert_partners(conn, parsed: dict) -> dict[str, int]:
    """
    parsed 구조를 받아 DB에 INSERT하고
    { 거래처명: id } 딕셔너리를 반환합니다.
    """
    name_to_id: dict[str, int] = {}
 
    with conn.cursor() as cur:
 
        # 1단계: parent 먼저 INSERT
        print(f"[1/2] parent 거래처 INSERT ({len(parsed['parents'])}개)")
        for name in sorted(parsed["parents"]):
            cur.execute("""
                INSERT INTO sales.res_partner (name, is_company, type)
                VALUES (%s, TRUE, 'contact')
                ON CONFLICT DO NOTHING
                RETURNING id
            """, (name,))
            row = cur.fetchone()
            if row:
                name_to_id[name] = row[0]
                print(f"  ✓ [{row[0]:>4}] {name}")
            else:
                # ON CONFLICT 로 스킵된 경우 → id 조회
                cur.execute(
                    "SELECT id FROM sales.res_partner WHERE name = %s", (name,)
                )
                name_to_id[name] = cur.fetchone()[0]
                print(f"  - (already exists) {name}")
 
        # 2단계: child INSERT (parent_id 세팅)
        print(f"\n[2/2] child 납품처 INSERT ({len(parsed['children'])}개)")
        for name, parent_name in sorted(parsed["children"].items()):
            parent_id = name_to_id.get(parent_name)
            cur.execute("""
                INSERT INTO sales.res_partner (name, is_company, type, parent_id)
                VALUES (%s, TRUE, 'delivery', %s)
                ON CONFLICT DO NOTHING
                RETURNING id
            """, (name, parent_id))
            row = cur.fetchone()
            if row:
                name_to_id[name] = row[0]
                print(f"  ✓ [{row[0]:>4}] {name}  (parent: {parent_name})")
            else:
                cur.execute(
                    "SELECT id FROM sales.res_partner WHERE name = %s", (name,)
                )
                name_to_id[name] = cur.fetchone()[0]
                print(f"  - (already exists) {name}")
 
    conn.commit()
    return name_to_id
 
 
def main():
    # ── 1. 엑셀에서 거래처명 수집 ──────────────────────────────
    print(f"엑셀 읽는 중: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
 
    raw_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[2]  # 대분류 컬럼
        if name:
            raw_names.add(name.strip())  # 앞뒤 공백 제거!
 
    print(f"고유 거래처명 {len(raw_names)}개 발견\n")
 
    # ── 2. parent/child 분석 ───────────────────────────────────
    parsed = parse_partners(raw_names)
    print(f"parent 거래처: {len(parsed['parents'])}개")
    print(f"child  납품처: {len(parsed['children'])}개\n")
 
    # ── 3. DB INSERT ───────────────────────────────────────────
    print("DB 연결 중...")
    conn = psycopg2.connect(**DB)
    try:
        name_to_id = insert_partners(conn, parsed)
        print(f"\n완료! 총 {len(name_to_id)}개 거래처 적재됨")
    except Exception as e:
        conn.rollback()
        print(f"\n에러 발생, rollback: {e}")
        raise
    finally:
        conn.close()
 
    return name_to_id
 
 
if __name__ == "__main__":
    main()
 