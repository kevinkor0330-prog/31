"""
load_product.py
DATA 시트를 읽어서 sales.product 테이블에 parent/child 구조로 INSERT합니다.

parent/child 규칙:
    괄호 안이 '숫자+배' 패턴  → child  (예: I-3(30배), 약품박스(ARC)60배)
    #숫자 패턴               → child  (예: EPO-55#2, MFZ63572201#92)
    그 외                   → parent (예: 약품박스(ARC), BIUP(수출용))

실행:
    pip install openpyxl psycopg2-binary
    python load_product.py
"""

import re
import psycopg2
from openpyxl import load_workbook
from datetime import date

# ── 설정 ──────────────────────────────────────────────────────
DB = {
    "host":     "158.247.208.195",
    "port":     5432,
    "dbname":   "db_hwaseong",
    "user":     "admin",
    "password": "1234",
}
EXCEL_PATH = "복사본_출하증_2023_원본_2025_.xlsm"
SHEET_NAME = "DATA"
# ─────────────────────────────────────────────────────────────


# 배율 변형 패턴: (30배), (33배), (60배) 등
RATIO_PATTERN = re.compile(r'^\d+배$')
# #숫자 패턴: EPO-55#2, 1208PAD#4
HASH_PATTERN  = re.compile(r'^(.+?)#(\d+)(.*)$')


def get_parent_base(code: str) -> str | None:
    """
    child 모델이면 parent 코드(base)를 반환.
    parent 모델이면 None 반환.
    """
    code = code.strip()

    # 1. #숫자 패턴: EPO-55#2 → base='EPO-55'
    m = HASH_PATTERN.match(code)
    if m:
        return m.group(1).strip()

    # 2. 괄호 패턴: 괄호 안이 '숫자+배' 이면 child
    m = re.match(r'^(.+?)[\(](.+?)[\)](.*)$', code)
    if m:
        base   = m.group(1).strip()
        inside = m.group(2).strip()
        suffix = m.group(3).strip()

        if RATIO_PATTERN.match(inside):
            # I-3(30배) → base='I-3'
            # 약품박스(ARC)30배 → inside='ARC' → NOT 배율, 아래 suffix 처리
            full_base = base + suffix  # suffix 있으면 붙임
            return full_base.strip()

        # 괄호 안이 배율이 아닌데 뒤에 배율이 붙는 경우
        # 약품박스(ARC)30배 → 괄호 통째로 base의 일부로 보고
        # 뒤의 '30배'는 child 구분자
        if suffix and RATIO_PATTERN.match(suffix):
            # 약품박스(ARC)30배 → base='약품박스(ARC)'
            return (base + '(' + inside + ')').strip()

    return None


def parse_products(rows: list) -> dict:
    """
    DATA 시트 rows를 받아 parent/child 구조 분석.
    반환: { code: { 'data': row_dict, 'parent_base': str|None } }
    """
    all_codes = {r['code'] for r in rows}
    result = {}

    for row in rows:
        code = row['code']
        base = get_parent_base(code)

        # base가 데이터에 없으면 → 자동생성 필요 (나중에 처리)
        result[code] = {
            'data':        row,
            'parent_base': base,
        }

    return result


def to_date(val):
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date()
    return None


def to_float(val):
    if val is None or val == '' or val == 0:
        return None
    try:
        # 쉼표 제거 후 변환
        return float(str(val).replace(',', ''))
    except:
        return None


def to_int(val):
    if val is None or val == '':
        return None
    try:
        return int(val)
    except:
        return None


def read_excel(path: str, sheet: str) -> list[dict]:
    """DATA 시트를 읽어서 딕셔너리 리스트로 반환."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    result = []
    for r in rows:
        code = r[0]
        if not code:
            continue
        code = str(code).strip()
        if not code:
            continue

        result.append({
            'code':          code,
            'category':      r[1],
            'cavity':        to_int(r[2]),
            'cycle_time':    to_int(r[3]),
            'weight_g':      to_float(r[4]),
            'ratio':         to_float(r[5]),
            'grade':         r[6],
            'pack_unit':     str(r[7]) if r[7] else None,
            'efficiency':    to_float(r[8]),
            'partner_name':  str(r[9]).strip() if r[9] else None,
            't_size_lo':     to_float(r[10]),
            't_size_up':     to_float(r[11]),
            'b_size_lo':     to_float(r[12]),
            'b_size_up':     to_float(r[13]),
            'base_price':    to_float(r[14]),
            'spec':          to_float(r[15]),
            'thickness':     to_float(r[16]),
            'strap':         str(r[17]) if r[17] else None,
            'color':         str(r[18]) if r[18] else None,
            'original_name': str(r[19]).strip() if r[19] else None,
            'mold_in_date':  to_date(r[20]),
            'mold_maker':    str(r[21]).strip() if r[21] else None,
            'model_sonata':  str(r[22]).strip() if r[22] else None,
            'mold_location': str(r[23]).strip() if r[23] else None,
            'base_ratio':    to_float(r[24]),
            'volume':        to_float(r[25]),
            'mold_out_date': to_date(r[26]),
            'mold_out_to':   str(r[27]).strip() if r[27] else None,
            'note':          str(r[28]).strip() if r[28] else None,
            'moisture_pct':  to_float(r[29]),
        })

    return result


INSERT_SQL = """
    INSERT INTO sales.product (
        parent_id, code, original_name, category,
        weight_g, ratio, grade, pack_unit, cavity, cycle_time, efficiency,
        spec, thickness, t_size_lo, t_size_up, b_size_lo, b_size_up,
        color, strap, base_price, partner_name,
        mold_maker, mold_location, mold_in_date,
        mold_out_date, mold_out_to, model_sonata,
        base_ratio, volume, moisture_pct, note
    ) VALUES (
        %(parent_id)s, %(code)s, %(original_name)s, %(category)s,
        %(weight_g)s, %(ratio)s, %(grade)s, %(pack_unit)s, %(cavity)s, %(cycle_time)s, %(efficiency)s,
        %(spec)s, %(thickness)s, %(t_size_lo)s, %(t_size_up)s, %(b_size_lo)s, %(b_size_up)s,
        %(color)s, %(strap)s, %(base_price)s, %(partner_name)s,
        %(mold_maker)s, %(mold_location)s, %(mold_in_date)s,
        %(mold_out_date)s, %(mold_out_to)s, %(model_sonata)s,
        %(base_ratio)s, %(volume)s, %(moisture_pct)s, %(note)s
    )
    ON CONFLICT DO NOTHING
    RETURNING id
"""


def insert_products(conn, parsed: dict) -> dict[str, int]:
    """
    1단계: parent_base 없는 것 (parent) 먼저 INSERT
    2단계: parent_base 있는 것 (child) INSERT
    """
    code_to_id: dict[str, int] = {}

    parents  = {c: v for c, v in parsed.items() if v['parent_base'] is None}
    children = {c: v for c, v in parsed.items() if v['parent_base'] is not None}

    print(f"[1/2] parent 모델 INSERT ({len(parents)}개)")
    with conn.cursor() as cur:
        for code, item in sorted(parents.items()):
            rec = {**item['data'], 'parent_id': None}
            cur.execute(INSERT_SQL, rec)
            row = cur.fetchone()
            if row:
                code_to_id[code] = row[0]
                print(f"  ✓ [{row[0]:>4}] {code}")
            else:
                cur.execute("SELECT id FROM sales.product WHERE code = %s", (code,))
                r = cur.fetchone()
                if r:
                    code_to_id[code] = r[0]
                    print(f"  - (exists) {code}")

    print(f"\n[2/2] child 모델 INSERT ({len(children)}개)")
    with conn.cursor() as cur:
        for code, item in sorted(children.items()):
            base       = item['parent_base']
            parent_id  = code_to_id.get(base)

            if parent_id is None:
                # parent가 DB에 없으면 먼저 조회
                cur.execute("SELECT id FROM sales.product WHERE code = %s", (base,))
                r = cur.fetchone()
                if r:
                    parent_id = r[0]
                    code_to_id[base] = parent_id
                else:
                    print(f"  ⚠ parent 없음, parent_id=NULL로 INSERT: {code} (base={base})")

            rec = {**item['data'], 'parent_id': parent_id}
            cur.execute(INSERT_SQL, rec)
            row = cur.fetchone()
            if row:
                code_to_id[code] = row[0]
                print(f"  ✓ [{row[0]:>4}] {code}  (parent: {base}, id={parent_id})")
            else:
                cur.execute("SELECT id FROM sales.product WHERE code = %s", (code,))
                r = cur.fetchone()
                if r:
                    code_to_id[code] = r[0]

    conn.commit()
    return code_to_id


def main():
    print(f"엑셀 읽는 중: {EXCEL_PATH} [{SHEET_NAME}]")
    rows = read_excel(EXCEL_PATH, SHEET_NAME)
    print(f"총 {len(rows)}개 모델 발견\n")

    parsed = parse_products(rows)

    parents  = sum(1 for v in parsed.values() if v['parent_base'] is None)
    children = sum(1 for v in parsed.values() if v['parent_base'] is not None)
    print(f"parent: {parents}개 / child: {children}개\n")

    print("DB 연결 중...")
    conn = psycopg2.connect(**DB)
    try:
        code_to_id = insert_products(conn, parsed)
        print(f"\n✓ 완료! 총 {len(code_to_id)}개 모델 적재")
    except Exception as e:
        conn.rollback()
        print(f"\n에러, rollback: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
