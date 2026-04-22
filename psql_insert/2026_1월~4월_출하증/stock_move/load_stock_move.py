"""
load_stock_move.py
당월자료관리 엑셀의 출하자료/화성출하 시트를 읽어서
sales.stock_move 테이블에 INSERT합니다.

실행:
    python load_stock_move.py
"""

import psycopg2
from openpyxl import load_workbook

# ── 설정 ──────────────────────────────────────────────────────
DB = {
    "host":     "158.247.208.195",
    "port":     5432,
    "dbname":   "db_hwaseong",
    "user":     "admin",
    "password": "1234",
}

#
EXCEL_FILES = [
    {"path": "당월자료관리26-4.xlsx", "sheets": ["출하자료", "화성출하"]},
]
# ─────────────────────────────────────────────────────────────


def load_partner_map(conn) -> dict[str, int]:
    """거래처명 → id 매핑"""
    with conn.cursor() as cur:
        cur.execute("SELECT id, name FROM sales.res_partner")
        return {name.strip(): pid for pid, name in cur.fetchall()}


def load_product_map(conn) -> dict[str, int]:
    """모델코드 → id 매핑"""
    with conn.cursor() as cur:
        cur.execute("SELECT id, code FROM sales.product")
        return {code.strip(): pid for pid, code in cur.fetchall()}


def read_sheet(path: str, sheet_name: str) -> list[dict]:
    """
    출하자료/화성출하 시트 읽기
    컬럼 순서: 순번, 일자, 거래처, 납품처, 모델, 생산수, 출하수, 차량No, ...
    """
    wb = load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ 시트 없음: {sheet_name}")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    result = []
    for r in rows:
        # 순번 없으면 스킵
        if not r[0]:
            continue

        # 일자
        raw_date = r[1]
        if raw_date is None:
            continue
        shipped_date = raw_date.date() if hasattr(raw_date, 'date') else raw_date

        # 거래처/납품처
        partner_name  = str(r[2]).strip() if r[2] else None
        delivery_name = str(r[3]).strip() if r[3] else None

        # 모델
        product_code = str(r[4]).strip() if r[4] else None

        # 출하수 (인덱스 6)
        shipped_qty = int(r[6]) if r[6] and str(r[6]).isdigit() else None
        if shipped_qty is None and r[6]:
            try:
                shipped_qty = int(float(str(r[6])))
            except:
                shipped_qty = None

        # 차량No (인덱스 7)
        vehicle = str(r[7]).strip() if r[7] else None

        result.append({
            "shipped_date":   shipped_date,
            "partner_name":   partner_name,
            "delivery_name":  delivery_name,
            "product_code":   product_code,
            "shipped_qty":    shipped_qty,
            "vehicle":        vehicle,
        })

    return result


def insert_stock_moves(conn, records: list[dict],
                       partner_map: dict, product_map: dict) -> tuple[int, int]:
    """
    INSERT 실행. (성공건수, 매칭실패건수) 반환
    """
    sql = """
        INSERT INTO sales.stock_move
            (shipped_date, partner_id, delivery_id, product_id, product_code, shipped_qty, vehicle)
        VALUES
            (%(shipped_date)s, %(partner_id)s, %(delivery_id)s, %(product_id)s,
             %(product_code)s, %(shipped_qty)s, %(vehicle)s)
    """

    success = 0
    no_match = 0

    with conn.cursor() as cur:
        for rec in records:
            partner_id  = partner_map.get(rec["partner_name"])
            delivery_id = partner_map.get(rec["delivery_name"])
            product_id  = product_map.get(rec["product_code"])

            if not partner_id:
                no_match += 1

            cur.execute(sql, {
                "shipped_date":  rec["shipped_date"],
                "partner_id":    partner_id,
                "delivery_id":   delivery_id,
                "product_id":    product_id,
                "product_code":  rec["product_code"],
                "shipped_qty":   rec["shipped_qty"],
                "vehicle":       rec["vehicle"],
            })
            success += 1

    conn.commit()
    return success, no_match


def main():
    print("DB 연결 중...")
    conn = psycopg2.connect(**DB)

    partner_map = load_partner_map(conn)
    product_map = load_product_map(conn)
    print(f"거래처 {len(partner_map)}개 / 모델 {len(product_map)}개 로드\n")

    total_success = 0
    total_no_match = 0

    for file_info in EXCEL_FILES:
        path = file_info["path"]
        print(f"=== {path} ===")

        for sheet in file_info["sheets"]:
            print(f"  [{sheet}] 읽는 중...")
            records = read_sheet(path, sheet)
            print(f"  [{sheet}] {len(records)}행 발견")

            if not records:
                continue

            try:
                success, no_match = insert_stock_moves(
                    conn, records, partner_map, product_map
                )
                total_success  += success
                total_no_match += no_match
                print(f"  [{sheet}] ✓ {success}건 INSERT / 거래처 미매칭: {no_match}건")
            except Exception as e:
                conn.rollback()
                print(f"  [{sheet}] 에러: {e}")
                raise

    print(f"\n✓ 완료! 총 {total_success}건 / 거래처 미매칭: {total_no_match}건")

    # 미매칭 거래처 확인
    if total_no_match > 0:
        print("\n--- 미매칭 거래처 확인 쿼리 ---")
        print("SELECT DISTINCT product_code, partner_id")
        print("FROM sales.stock_move WHERE partner_id IS NULL;")

    conn.close()


if __name__ == "__main__":
    main()
