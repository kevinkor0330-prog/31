"""
거래처 계층 구조 테이블 생성
- 슬래시(/) 앞 = Parent (기업 본사/주관사)
- 슬래시(/) 뒤 = Child (지점/납품처/하위거래처)
- 단독 행 = 슬래시 없는 독립 거래처 (Parent = None)
- 슬래시 패턴에만 등장하는 Parent → 자동으로 Parent 행 생성
출력: 하나의 Master 테이블에서 parent_id로 계층 조회 가능
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 0. 파일 로드
# ─────────────────────────────────────────────
FILE_PATH = "연간계휙표_단가.xlsx"

wb = openpyxl.load_workbook(FILE_PATH)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

columns = [
    "key", "영업담당자", "거래처명", "삼일모델명", "업체모델명", "CODE_No",
    "단가", "합의중량_g", "구단가",
    "구단가_1", "구단가_2", "구단가_3", "구단가_4",
    "대표자", "사업자번호", "주소", "업태", "종목", "납품주소",
    "마감담당자_성명", "마감담당자_이메일", "마감담당자_연락처",
    "실무담당자_성명", "실무담당자_이메일", "실무담당자_연락처",
    "운임", "결제조건_일", "제품군", "비고"
]

data = [row for row in rows[3:] if any(row)]
df_raw = pd.DataFrame(data, columns=columns).drop(columns=["key"])
df_raw["거래처명"] = df_raw["거래처명"].astype(str).str.strip()
df_raw = df_raw[df_raw["거래처명"].notna() & (df_raw["거래처명"] != "None")].reset_index(drop=True)

# 거래처별 대표 정보 (첫 번째 등장 행 기준)
META_COLS = [
    "거래처명", "대표자", "사업자번호", "주소", "업태", "종목", "납품주소",
    "마감담당자_성명", "마감담당자_이메일", "마감담당자_연락처",
    "실무담당자_성명", "실무담당자_이메일", "실무담당자_연락처",
    "운임", "결제조건_일"
]
df_meta = df_raw[META_COLS].drop_duplicates(subset=["거래처명"]).set_index("거래처명")


# ─────────────────────────────────────────────
# 1. 고유 거래처명 수집 & 슬래시 분석
# ─────────────────────────────────────────────
all_clients = set(df_raw["거래처명"].unique())

slash_clients = {c for c in all_clients if "/" in c}
standalone_clients = {c for c in all_clients if "/" not in c}

# 슬래시 패턴에서 parent명 추출
parent_names_from_slash = {c.split("/")[0].strip() for c in slash_clients}

# 슬래시에만 있고 단독 행이 없는 parent → 자동 생성 필요
auto_create_parents = parent_names_from_slash - standalone_clients

print(f"전체 고유 거래처: {len(all_clients)}개")
print(f"  슬래시(/) 패턴: {len(slash_clients)}개")
print(f"  단독 행: {len(standalone_clients)}개")
print(f"  자동 생성할 Parent 행: {len(auto_create_parents)}개")


# ─────────────────────────────────────────────
# 2. Master 테이블 구성
# ─────────────────────────────────────────────
records = []

# ── Step A: 단독 행 거래처 등록 ──────────────
for name in sorted(standalone_clients):
    is_parent = name in parent_names_from_slash  # 슬래시 child의 parent이기도 함
    meta = df_meta.loc[name].to_dict() if name in df_meta.index else {}
    records.append({
        "거래처명": name,
        "parent_명": None,          # 단독이므로 부모 없음
        "node_type": "parent" if is_parent else "독립",
        # 메타
        "대표자": meta.get("대표자"),
        "사업자번호": meta.get("사업자번호"),
        "주소": meta.get("주소"),
        "업태": meta.get("업태"),
        "종목": meta.get("종목"),
        "납품주소": meta.get("납품주소"),
        "마감담당자_성명": meta.get("마감담당자_성명"),
        "마감담당자_이메일": meta.get("마감담당자_이메일"),
        "마감담당자_연락처": meta.get("마감담당자_연락처"),
        "실무담당자_성명": meta.get("실무담당자_성명"),
        "실무담당자_이메일": meta.get("실무담당자_이메일"),
        "실무담당자_연락처": meta.get("실무담당자_연락처"),
        "운임": meta.get("운임"),
        "결제조건_일": meta.get("결제조건_일"),
    })

# ── Step B: 자동 생성 Parent 행 등록 ─────────
for name in sorted(auto_create_parents):
    records.append({
        "거래처명": name,
        "parent_명": None,
        "node_type": "parent",
        "대표자": None, "사업자번호": None, "주소": None,
        "업태": None, "종목": None, "납품주소": None,
        "마감담당자_성명": None, "마감담당자_이메일": None, "마감담당자_연락처": None,
        "실무담당자_성명": None, "실무담당자_이메일": None, "실무담당자_연락처": None,
        "운임": None, "결제조건_일": None,
    })

# ── Step C: Child 행 등록 ─────────────────────
for name in sorted(slash_clients):
    parent_name = name.split("/")[0].strip()
    child_label = name.split("/", 1)[1].strip()
    meta = df_meta.loc[name].to_dict() if name in df_meta.index else {}
    records.append({
        "거래처명": name,
        "parent_명": parent_name,
        "node_type": "child",
        "대표자": meta.get("대표자"),
        "사업자번호": meta.get("사업자번호"),
        "주소": meta.get("주소"),
        "업태": meta.get("업태"),
        "종목": meta.get("종목"),
        "납품주소": meta.get("납품주소"),
        "마감담당자_성명": meta.get("마감담당자_성명"),
        "마감담당자_이메일": meta.get("마감담당자_이메일"),
        "마감담당자_연락처": meta.get("마감담당자_연락처"),
        "실무담당자_성명": meta.get("실무담당자_성명"),
        "실무담당자_이메일": meta.get("실무담당자_이메일"),
        "실무담당자_연락처": meta.get("실무담당자_연락처"),
        "운임": meta.get("운임"),
        "결제조건_일": meta.get("결제조건_일"),
    })

df_master = pd.DataFrame(records)

# ── 숫자 ID 부여 ──────────────────────────────
df_master = df_master.reset_index(drop=True)
df_master.insert(0, "거래처_ID", range(1, len(df_master) + 1))

# ── parent_id FK 연결 ─────────────────────────
name_to_id = df_master.set_index("거래처명")["거래처_ID"].to_dict()
df_master["parent_id"] = df_master["parent_명"].map(name_to_id)

# parent_id를 두 번째 컬럼으로 이동
cols = df_master.columns.tolist()
cols.remove("parent_id")
cols.insert(2, "parent_id")
df_master = df_master[cols]

# node_type 순서 정렬: parent/독립 먼저, child 뒤
type_order = {"parent": 0, "독립": 1, "child": 2}
df_master["_sort"] = df_master["node_type"].map(type_order)
df_master = df_master.sort_values(["_sort", "거래처_ID"]).drop(columns=["_sort"]).reset_index(drop=True)
df_master["거래처_ID"] = range(1, len(df_master) + 1)
# parent_id 재연결 (정렬 후 ID 바뀌었으므로)
name_to_id = df_master.set_index("거래처명")["거래처_ID"].to_dict()
df_master["parent_id"] = df_master["parent_명"].map(name_to_id)


# ─────────────────────────────────────────────
# 3. Detail 테이블 (FK 연결)
# ─────────────────────────────────────────────
df_raw["거래처_ID"] = df_raw["거래처명"].map(name_to_id)

detail_cols = [
    "거래처_ID", "거래처명", "영업담당자",
    "삼일모델명", "업체모델명", "CODE_No",
    "단가", "합의중량_g", "구단가",
    "구단가_1", "구단가_2", "구단가_3", "구단가_4",
    "제품군", "비고"
]
df_detail = df_raw[detail_cols].reset_index(drop=True)
df_detail.insert(0, "거래내역_ID", range(1, len(df_detail) + 1))

for col in ["단가", "합의중량_g", "구단가", "구단가_1", "구단가_2", "구단가_3", "구단가_4"]:
    df_detail[col] = pd.to_numeric(df_detail[col], errors="coerce")


# ─────────────────────────────────────────────
# 4. 결과 요약
# ─────────────────────────────────────────────
parents   = df_master[df_master["node_type"] == "parent"]
children  = df_master[df_master["node_type"] == "child"]
standalone= df_master[df_master["node_type"] == "독립"]

print(f"\n{'='*55}")
print("【Master 테이블 계층 구조 요약】")
print(f"  Parent  (본사/주관사)  : {len(parents):>4}개")
print(f"  Child   (지점/납품처)  : {len(children):>4}개")
print(f"  독립    (단독 거래처)  : {len(standalone):>4}개")
print(f"  전체                  : {len(df_master):>4}개")
print(f"  Detail  (거래내역)    : {len(df_detail):>4}개")
print(f"{'='*55}")

print("\n【Parent → Child 샘플】")
for p_name in ["청호나이스", "CJ대한통운", "동원홈푸드", "한국파렛트풀", "컬리"]:
    if p_name not in name_to_id:
        continue
    pid = name_to_id[p_name]
    kids = df_master[df_master["parent_id"] == pid][["거래처_ID", "거래처명"]].values
    print(f"  [{pid}] {p_name}")
    for kid_id, kid_name in kids[:5]:
        print(f"       └─ [{kid_id}] {kid_name}")
    if len(kids) > 5:
        print(f"       └─ ... 외 {len(kids)-5}개")


# ─────────────────────────────────────────────
# 5. Excel 저장
# ─────────────────────────────────────────────
OUTPUT_PATH = "SQL_계층구조_거래처DB.xlsx"

# 색상 팔레트
COLOR = {
    "header_bg": "1B2A3B",
    "parent_bg": "D6EAF8",
    "parent_font": "1A5276",
    "child_bg":  "D5F5E3",
    "child_font": "1E8449",
    "indep_bg":  "FEF9E7",
    "indep_font": "7D6608",
    "alt_row": "F4F6F7",
}

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    df_master.to_excel(writer, sheet_name="Master_계층구조", index=False)
    df_detail.to_excel(writer, sheet_name="Detail_거래내역", index=False)

    # ── Master 시트 서식 ──────────────────────
    ws_m = writer.sheets["Master_계층구조"]
    header_fill = PatternFill("solid", start_color=COLOR["header_bg"], end_color=COLOR["header_bg"])
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws_m[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    node_col_idx = df_master.columns.get_loc("node_type") + 1

    for row_idx, row in enumerate(ws_m.iter_rows(min_row=2), start=2):
        node_val = ws_m.cell(row=row_idx, column=node_col_idx).value

        if node_val == "parent":
            bg, fc, bold = COLOR["parent_bg"], COLOR["parent_font"], True
        elif node_val == "child":
            bg, fc, bold = COLOR["child_bg"], COLOR["child_font"], False
        else:  # 독립
            bg, fc, bold = COLOR["indep_bg"], COLOR["indep_font"], False

        fill = PatternFill("solid", start_color=bg, end_color=bg)
        font = Font(color=fc, bold=bold, name="Arial", size=9)
        for cell in row:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center")

    # 열 너비 조정
    for col in ws_m.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_m.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    ws_m.freeze_panes = "A2"
    ws_m.auto_filter.ref = ws_m.dimensions

    # ── Detail 시트 서식 ──────────────────────
    ws_d = writer.sheets["Detail_거래내역"]
    for cell in ws_d[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill("solid", start_color=COLOR["alt_row"], end_color=COLOR["alt_row"])
    for row_idx, row in enumerate(ws_d.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill
        for cell in row:
            cell.font = Font(name="Arial", size=9)

    for col in ws_d.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_d.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws_d.freeze_panes = "A2"
    ws_d.auto_filter.ref = ws_d.dimensions

print(f"\n✅ 저장 완료: {OUTPUT_PATH}")

# ─────────────────────────────────────────────
# 6. SQL 참고 DDL
# ─────────────────────────────────────────────
print("""
【SQL CREATE TABLE】

CREATE TABLE master_client (
    거래처_ID      INT PRIMARY KEY,
    거래처명       VARCHAR(150) NOT NULL,
    parent_id      INT NULL,                          -- self-referencing FK
    parent_명      VARCHAR(150) NULL,
    node_type      ENUM('parent','child','독립') NOT NULL,
    대표자         VARCHAR(50),
    사업자번호     VARCHAR(20),
    주소           VARCHAR(200),
    업태           VARCHAR(100),
    종목           VARCHAR(100),
    납품주소       VARCHAR(200),
    마감담당자_성명    VARCHAR(50),
    마감담당자_이메일  VARCHAR(100),
    마감담당자_연락처  VARCHAR(20),
    실무담당자_성명    VARCHAR(50),
    실무담당자_이메일  VARCHAR(100),
    실무담당자_연락처  VARCHAR(20),
    운임            VARCHAR(50),
    결제조건_일     INT,
    FOREIGN KEY (parent_id) REFERENCES master_client(거래처_ID)
);

-- 계층 조회 예시
SELECT
    p.거래처_ID   AS parent_id,
    p.거래처명    AS parent_명,
    c.거래처_ID   AS child_id,
    c.거래처명    AS child_명
FROM master_client p
LEFT JOIN master_client c ON c.parent_id = p.거래처_ID
WHERE p.node_type = 'parent'
ORDER BY p.거래처명, c.거래처명;
""")
